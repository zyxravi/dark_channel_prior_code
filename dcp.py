import cv2
import numpy as np
from skimage.metrics import structural_similarity as ssim
import openpyxl

def dark_channel_prior(img, patch_size=15):
    b, g, r = cv2.split(img)
    min_channel = cv2.min(cv2.min(r, g), b)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (patch_size, patch_size))
    dark_channel = cv2.erode(min_channel, kernel)
    return dark_channel


def atmospheric_light(im, dc, percentage=0.001):
    # Estimate the atmospheric light of the image for each channel
    A = np.zeros(3)
    for i in range(3):  # Iterate over R, G, B channels
        A[i] = np.max(im[:, :, i][np.where(dc == np.max(dc))])
    return A


def transmission_estimate(img, A, omega=0.95, patch_size=15):
    normalized_image = img / A
    dark_channel = dark_channel_prior(normalized_image, patch_size)
    transmission = 1 - omega * dark_channel
    return transmission


def recover_radiance(img, A, transmission, Rho):
    recovered_radiance = np.zeros_like(img, dtype=np.float64)
    for i in range(3):
        recovered_radiance[:, :, i] = (img[:, :, i] - A[i]) / np.maximum(transmission, 0.1) + \
                                      A[i] - Rho
    return np.clip(recovered_radiance, 0, 255).astype(np.uint8)


rho = 0.75


def dehaze(img, omega=0.95, percent=0.001, patch_size=15):
    dark_channel = dark_channel_prior(img, patch_size)
    A = atmospheric_light(img, dark_channel, percent)
    transmission = transmission_estimate(img, A, omega, patch_size)
    recovered_image = recover_radiance(img, A, transmission, rho)
    return recovered_image

def calculation(raw_image, defogged):
    if raw_image.shape == defogged.shape:
        gray1 = cv2.cvtColor(raw_image, cv2.COLOR_BGR2GRAY)
        gray2 = cv2.cvtColor(defogged, cv2.COLOR_BGR2GRAY)
        # MSE stands for mean square error
        MSE = np.sum((raw_image.astype("float32") - defogged.astype("float32")) ** 2)
        MSE /= float(raw_image.shape[0] * defogged.shape[1])
        s = ssim(gray1, gray2)
        PSNR = cv2.PSNR(raw_image, defogged)
        return MSE, s, PSNR
    else:
        print("hazy and defogged image have the different dimensions")



wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row=1, column=1)
c1.value = "Serial No."
c2 = sheet.cell(row=1, column=2)
c2.value = "MSE"
c3 = sheet.cell(row=1, column=3)
c3.value = "SSIM"

c3 = sheet.cell(row=1, column=4)
c3.value = "PSNR"

for i in range(100):
    image = cv2.imread(f'hazy_image2/frame1_{i:04d}.jpg')
    defogged_image = dehaze(image)
    output_image = f"dehazed_image/{frame1_{i:04d}.jpg"
    cv2.imwrite(output_image, defogged_image)
    MSE, SSIM, PSNR = calculation(image, defogged_image)
    c1 = sheet.cell(row=i + 2, column=1)
    c1.value = i + 1
    c2 = sheet.cell(row=i + 2, column=2)
    c2.value = MSE
    c3 = sheet.cell(row=i + 2, column=3)
    c3.value = SSIM
    c4 = sheet.cell(row=i + 2, column=4)
    c4.value = PSNR
wb.save("calculation.xlsx")
cv2.destroyAllWindows()
